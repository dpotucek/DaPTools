#!/usr/bin/env python3
"""
Generate a compilable LaTeX document with a table for diluting alcohol to 40%.

Rows are source alcohol concentrations, columns are source alcohol amounts,
and each cell contains the amount of water to add in millilitres.
"""

import argparse
from pathlib import Path


CONCENTRATIONS = [50, 55, 60, 65, 70, 75, 80, 85, 90]
AMOUNTS_ML = [10, 20, 25, 30, 40, 50, 100, 200, 250, 500, 750]
TARGET_VOLUMES_L = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
TARGET_CONCENTRATION = 40


def water_to_add_ml(amount_ml: float, source_percent: float,
                    target_percent: float = TARGET_CONCENTRATION) -> float:
    """Return millilitres of water needed to dilute to target_percent."""
    if amount_ml <= 0:
        raise ValueError("Amount must be positive")
    if source_percent <= 0 or source_percent > 100:
        raise ValueError("Source concentration must be between 0 and 100 %")
    if target_percent <= 0 or target_percent > 100:
        raise ValueError("Target concentration must be between 0 and 100 %")
    if target_percent >= source_percent:
        raise ValueError("Target concentration must be lower than source concentration")

    return amount_ml * (source_percent / target_percent - 1)


def mix_for_target_volume_ml(final_volume_ml: float, source_percent: float,
                             target_percent: float = TARGET_CONCENTRATION) -> tuple[float, float]:
    """Return source alcohol and water amounts needed for final_volume_ml."""
    if final_volume_ml <= 0:
        raise ValueError("Final volume must be positive")
    if source_percent <= 0 or source_percent > 100:
        raise ValueError("Source concentration must be between 0 and 100 %")
    if target_percent <= 0 or target_percent > 100:
        raise ValueError("Target concentration must be between 0 and 100 %")
    if target_percent >= source_percent:
        raise ValueError("Target concentration must be lower than source concentration")

    alcohol_ml = final_volume_ml * target_percent / source_percent
    water_ml = final_volume_ml - alcohol_ml
    return alcohol_ml, water_ml


def format_ml(value: float) -> str:
    """Format millilitres compactly for the table."""
    rounded = round(value, 1)
    if rounded.is_integer():
        return str(int(rounded))
    return f"{rounded:.1f}"


def format_liters(value: float) -> str:
    """Format litres for table headers."""
    return f"{value:.1f} l"


def format_mix_cell(alcohol_ml: float, water_ml: float) -> str:
    """Format alcohol and water amounts compactly for the table."""
    return rf"{format_ml(alcohol_ml)} + {format_ml(water_ml)}"


def generate_water_latex_table() -> str:
    column_spec = "|" + "|".join(["r"] * (len(AMOUNTS_ML) + 1)) + "|"
    header = r"\textbf{Koncentrace}"
    amount_headers = [rf"\textbf{{{amount} ml}}" for amount in AMOUNTS_ML]

    lines = [
        rf"\begin{{center}}\Large\textbf{{Množství vody v ml pro ředění alkoholu na {TARGET_CONCENTRATION}\%}}\end{{center}}",
        r"\resizebox{\textwidth}{!}{%",
        rf"\begin{{tabular}}{{{column_spec}}}",
        r"\hline",
        " & ".join([header, *amount_headers]) + r" \\",
        r"\hline",
    ]

    for concentration in CONCENTRATIONS:
        values = [
            format_ml(water_to_add_ml(amount, concentration))
            for amount in AMOUNTS_ML
        ]
        lines.append(" & ".join([rf"\textbf{{{concentration}\%}}", *values]) + r" \\")
        lines.append(r"\hline")

    lines.extend([
        r"\end{tabular}",
        r"}",
    ])
    return "\n".join(lines)


def generate_target_volume_latex_table() -> str:
    column_spec = "|" + "|".join(["r"] * (len(TARGET_VOLUMES_L) + 1)) + "|"
    header = r"\textbf{Koncentrace}"
    volume_headers = [rf"\textbf{{{format_liters(volume)}}}" for volume in TARGET_VOLUMES_L]

    lines = [
        rf"\begin{{center}}\Large\textbf{{Množství alkoholu a vody v ml pro výsledný objem při ředění na {TARGET_CONCENTRATION}\%}}\end{{center}}",
        r"\resizebox{\textwidth}{!}{%",
        rf"\begin{{tabular}}{{{column_spec}}}",
        r"\hline",
        " & ".join([header, *volume_headers]) + r" \\",
        r"\hline",
    ]

    for concentration in CONCENTRATIONS:
        values = [
            format_mix_cell(*mix_for_target_volume_ml(volume * 1000, concentration))
            for volume in TARGET_VOLUMES_L
        ]
        lines.append(" & ".join([rf"\textbf{{{concentration}\%}}", *values]) + r" \\")
        lines.append(r"\hline")

    lines.extend([
        r"\end{tabular}",
        r"}",
    ])
    return "\n".join(lines)


def generate_latex_document() -> str:
    lines = [
        r"\documentclass[a4paper,10pt]{article}",
        r"\usepackage[landscape, margin=0.7cm]{geometry}",
        r"\usepackage{graphicx}",
        r"\pagestyle{empty}",
        r"\begin{document}",
        generate_water_latex_table(),
        r"\clearpage",
        generate_target_volume_latex_table(),
        r"\end{document}",
    ]
    return "\n".join(lines)


def write_excel_workbook(path: Path) -> None:
    """Write both dilution tables into separate sheets of one XLSX workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel output requires the openpyxl package") from exc

    workbook = Workbook()
    water_sheet = workbook.active
    water_sheet.title = "Voda k pridani"
    target_sheet = workbook.create_sheet("Cilovy objem")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    water_sheet.append(["Koncentrace", *[f"{amount} ml" for amount in AMOUNTS_ML]])
    for concentration in CONCENTRATIONS:
        water_sheet.append([
            f"{concentration}%",
            *[round(water_to_add_ml(amount, concentration), 1) for amount in AMOUNTS_ML],
        ])

    target_sheet.append(["Koncentrace", *[format_liters(volume) for volume in TARGET_VOLUMES_L]])
    for concentration in CONCENTRATIONS:
        target_sheet.append([
            f"{concentration}%",
            *[
                format_mix_cell(*mix_for_target_volume_ml(volume * 1000, concentration))
                for volume in TARGET_VOLUMES_L
            ],
        ])

    for sheet in (water_sheet, target_sheet):
        sheet.freeze_panes = "B2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for cell in sheet["A"]:
            cell.font = header_font
            cell.fill = header_fill
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 18))

    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a compilable LaTeX document with water amounts for dilution to 40% alcohol."
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        help="Optional path where the generated LaTeX document should be written.",
    )
    parser.add_argument(
        "-x", "--excel-output",
        type=Path,
        help="Optional path where the generated XLSX workbook should be written.",
    )
    args = parser.parse_args()

    latex = generate_latex_document()
    if args.output:
        args.output.write_text(latex + "\n", encoding="utf-8")
    else:
        print(latex)

    if args.excel_output:
        write_excel_workbook(args.excel_output)


if __name__ == "__main__":
    main()
